from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Avg, Q
from django.utils import timezone
from datetime import datetime, timedelta
from .models import Survey, Response, Customer, SurveySummary
import json
import csv
import openpyxl
from openpyxl import Workbook


def dashboard(request):
    """Dashboard principal com gráficos e estatísticas"""
    # Estatísticas gerais
    total_surveys = Survey.objects.count()
    total_responses = Response.objects.count()
    total_customers = Customer.objects.count()
    
    # Respostas dos últimos 30 dias
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_responses = Response.objects.filter(submitted_at__gte=thirty_days_ago).count()
    
    # Distribuição NPS
    nps_distribution = Response.objects.filter(
        score__isnull=False,
        survey__survey_type='nps'
    ).values('score').annotate(count=Count('score')).order_by('score')
    
    # Categorias NPS
    promoters = Response.objects.filter(
        score__gte=9,
        survey__survey_type='nps'
    ).count()
    passives = Response.objects.filter(
        score__in=[7, 8],
        survey__survey_type='nps'
    ).count()
    detractors = Response.objects.filter(
        score__lte=6,
        survey__survey_type='nps'
    ).count()
    
    # Calcular NPS Score
    total_nps_responses = promoters + passives + detractors
    nps_score = ((promoters - detractors) / total_nps_responses * 100) if total_nps_responses > 0 else 0
    
    # Tendência temporal (últimos 7 dias)
    trend_data = []
    for i in range(7):
        date = timezone.now() - timedelta(days=i)
        day_responses = Response.objects.filter(
            submitted_at__date=date.date(),
            survey__survey_type='nps'
        )
        day_promoters = day_responses.filter(score__gte=9).count()
        day_detractors = day_responses.filter(score__lte=6).count()
        day_total = day_responses.count()
        day_nps = ((day_promoters - day_detractors) / day_total * 100) if day_total > 0 else 0
        
        trend_data.append({
            'date': date.strftime('%Y-%m-%d'),
            'nps': round(day_nps, 1),
            'responses': day_total
        })
    
    trend_data.reverse()
    
    context = {
        'total_surveys': total_surveys,
        'total_responses': total_responses,
        'total_customers': total_customers,
        'recent_responses': recent_responses,
        'nps_distribution': list(nps_distribution),
        'promoters': promoters,
        'passives': passives,
        'detractors': detractors,
        'nps_score': round(nps_score, 1),
        'trend_data': trend_data,
    }
    
    return render(request, 'dashboard/dashboard.html', context)


def surveys_list(request):
    """Lista todas as pesquisas"""
    surveys = Survey.objects.all().order_by('-created_at')
    return render(request, 'dashboard/surveys_list.html', {'surveys': surveys})


def responses_list(request):
    """Lista todas as respostas com filtros"""
    responses = Response.objects.select_related('customer', 'survey').all().order_by('-submitted_at')
    
    # Filtros
    survey_id = request.GET.get('survey')
    customer_id = request.GET.get('customer')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if survey_id:
        responses = responses.filter(survey_id=survey_id)
    if customer_id:
        responses = responses.filter(customer_id=customer_id)
    if date_from:
        responses = responses.filter(submitted_at__date__gte=date_from)
    if date_to:
        responses = responses.filter(submitted_at__date__lte=date_to)
    
    # Opções para filtros
    surveys = Survey.objects.all()
    customers = Customer.objects.all()
    
    context = {
        'responses': responses,
        'surveys': surveys,
        'customers': customers,
        'filters': {
            'survey_id': survey_id,
            'customer_id': customer_id,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    
    return render(request, 'dashboard/responses_list.html', context)


def export_responses_csv(request):
    """Exporta respostas para CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="respostas.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Cliente', 'Email', 'Empresa', 'Pesquisa', 'Tipo', 'Nota', 'Comentário', 'Data'])
    
    responses = Response.objects.select_related('customer', 'survey').all()
    for resp in responses:
        writer.writerow([
            resp.customer.name,
            resp.customer.email,
            resp.customer.company,
            resp.survey.title,
            resp.survey.get_survey_type_display(),
            resp.score or '',
            resp.comment,
            resp.submitted_at.strftime('%d/%m/%Y %H:%M')
        ])
    
    return response


def export_responses_excel(request):
    """Exporta respostas para Excel"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="respostas.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Respostas"
    
    # Cabeçalhos
    headers = ['Cliente', 'Email', 'Empresa', 'Pesquisa', 'Tipo', 'Nota', 'Comentário', 'Data']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Dados
    responses = Response.objects.select_related('customer', 'survey').all()
    for row, resp in enumerate(responses, 2):
        ws.cell(row=row, column=1, value=resp.customer.name)
        ws.cell(row=row, column=2, value=resp.customer.email)
        ws.cell(row=row, column=3, value=resp.customer.company)
        ws.cell(row=row, column=4, value=resp.survey.title)
        ws.cell(row=row, column=5, value=resp.survey.get_survey_type_display())
        ws.cell(row=row, column=6, value=resp.score or '')
        ws.cell(row=row, column=7, value=resp.comment)
        ws.cell(row=row, column=8, value=resp.submitted_at.strftime('%d/%m/%Y %H:%M'))
    
    wb.save(response)
    return response


def api_nps_data(request):
    """API para dados NPS (usado pelos gráficos)"""
    # Distribuição de notas NPS
    distribution = Response.objects.filter(
        score__isnull=False,
        survey__survey_type='nps'
    ).values('score').annotate(count=Count('score')).order_by('score')
    
    # Categorias NPS
    promoters = Response.objects.filter(score__gte=9, survey__survey_type='nps').count()
    passives = Response.objects.filter(score__in=[7, 8], survey__survey_type='nps').count()
    detractors = Response.objects.filter(score__lte=6, survey__survey_type='nps').count()
    
    # Tendência temporal
    trend_data = []
    for i in range(30):  # Últimos 30 dias
        date = timezone.now() - timedelta(days=i)
        day_responses = Response.objects.filter(
            submitted_at__date=date.date(),
            survey__survey_type='nps'
        )
        day_promoters = day_responses.filter(score__gte=9).count()
        day_detractors = day_responses.filter(score__lte=6).count()
        day_total = day_responses.count()
        day_nps = ((day_promoters - day_detractors) / day_total * 100) if day_total > 0 else 0
        
        trend_data.append({
            'date': date.strftime('%Y-%m-%d'),
            'nps': round(day_nps, 1),
            'responses': day_total
        })
    
    trend_data.reverse()
    
    data = {
        'distribution': list(distribution),
        'categories': {
            'promoters': promoters,
            'passives': passives,
            'detractors': detractors
        },
        'trend': trend_data
    }
    
    return JsonResponse(data)