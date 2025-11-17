"""
Views para dashboard NPS e exportação de relatórios
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Count, Q
from django.core.serializers.json import DjangoJSONEncoder
from datetime import datetime, timedelta
from decimal import Decimal
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import Survey, SurveyResponse, NPSResult, Answer, Question
from .utils import (
    calculate_and_save_nps, get_nps_summary, 
    get_nps_trend_data, calculate_nps_from_responses
)


@login_required
def nps_dashboard(request, survey_id=None):
    """Dashboard principal de NPS"""
    if survey_id:
        survey = get_object_or_404(Survey, id=survey_id)
        surveys = [survey]
    else:
        surveys = Survey.objects.filter(is_active=True).order_by('-created_at')
        survey = surveys.first() if surveys.exists() else None
    
    # Resumo geral
    summary = get_nps_summary(survey=survey)
    
    # Dados de tendência (últimos 6 meses)
    trend_data = []
    trend_data_json = json.dumps([])
    if survey:
        trend_data = get_nps_trend_data(survey, months=6)
        # Serializar datas para JSON
        trend_data_json = json.dumps(trend_data, cls=DjangoJSONEncoder)
    
    # Últimos resultados NPS calculados
    recent_results = []
    if survey:
        recent_results = NPSResult.objects.filter(survey=survey).order_by('-calculated_at')[:10]
    
    # Respostas recentes
    recent_responses = []
    if survey:
        recent_responses = SurveyResponse.objects.filter(survey=survey).order_by('-submitted_at')[:10]
    
    context = {
        'survey': survey,
        'surveys': surveys,
        'summary': summary,
        'trend_data': trend_data,
        'trend_data_json': trend_data_json,
        'recent_results': recent_results,
        'recent_responses': recent_responses,
    }
    
    return render(request, 'surveys/nps_dashboard.html', context)


@login_required
def nps_calculate(request, survey_id):
    """Calcula NPS para um período específico"""
    survey = get_object_or_404(Survey, id=survey_id)
    
    if request.method == 'POST':
        period_start_str = request.POST.get('period_start')
        period_end_str = request.POST.get('period_end')
        force_recalculate = request.POST.get('force_recalculate') == 'on'
        
        try:
            period_start = datetime.strptime(period_start_str, '%Y-%m-%d').date()
            period_end = datetime.strptime(period_end_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Formato de data inválido'}, status=400)
        
        nps_result = calculate_and_save_nps(
            survey, 
            period_start, 
            period_end, 
            force_recalculate=force_recalculate
        )
        
        if nps_result:
            return JsonResponse({
                'success': True,
                'result': {
                    'id': nps_result.id,
                    'nps_score': str(nps_result.nps_score),
                    'total_responses': nps_result.total_responses,
                    'promoters': nps_result.promoters,
                    'passives': nps_result.passives,
                    'detractors': nps_result.detractors,
                    'period_start': nps_result.period_start.strftime('%Y-%m-%d'),
                    'period_end': nps_result.period_end.strftime('%Y-%m-%d'),
                }
            })
        else:
            return JsonResponse({
                'error': 'Nenhuma resposta NPS encontrada para este período'
            }, status=400)
    
    return JsonResponse({'error': 'Método não permitido'}, status=405)


@login_required
def nps_api_data(request, survey_id):
    """API para obter dados NPS em JSON (para gráficos)"""
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Parâmetros
    months = int(request.GET.get('months', 6))
    
    # Dados de tendência
    trend_data = get_nps_trend_data(survey, months=months)
    
    # Resumo atual
    summary = get_nps_summary(survey=survey)
    
    return JsonResponse({
        'trend': trend_data,
        'summary': {
            'nps_score': float(summary['nps_score']),
            'total_responses': summary['total_responses'],
            'promoters': summary['promoters'],
            'passives': summary['passives'],
            'detractors': summary['detractors'],
            'promoter_percentage': summary['promoter_percentage'],
            'passive_percentage': summary['passive_percentage'],
            'detractor_percentage': summary['detractor_percentage'],
        }
    })


@login_required
def export_nps_excel(request, survey_id):
    """Exporta relatório NPS em Excel"""
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Criar workbook
    wb = Workbook()
    
    # Planilha 1: Resumo
    ws_summary = wb.active
    ws_summary.title = "Resumo NPS"
    
    # Cabeçalhos
    headers = ['Período', 'Total Respostas', 'Promotores', 'Neutros', 'Detratores', 'NPS Score']
    ws_summary.append(headers)
    
    # Estilizar cabeçalhos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Dados dos resultados NPS
    results = NPSResult.objects.filter(survey=survey).order_by('-period_start', '-period_end')
    for result in results:
        ws_summary.append([
            f"{result.period_start.strftime('%d/%m/%Y')} a {result.period_end.strftime('%d/%m/%Y')}",
            result.total_responses,
            result.promoters,
            result.passives,
            result.detractors,
            float(result.nps_score)
        ])
    
    # Ajustar largura das colunas
    for col_num, column in enumerate(ws_summary.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Planilha 2: Respostas Detalhadas
    ws_responses = wb.create_sheet("Respostas Detalhadas")
    
    # Cabeçalhos
    headers_responses = ['Data', 'Email', 'Pergunta', 'Resposta NPS', 'Comentário']
    ws_responses.append(headers_responses)
    
    # Estilizar cabeçalhos
    for col_num, header in enumerate(headers_responses, 1):
        cell = ws_responses.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Dados das respostas
    responses = SurveyResponse.objects.filter(survey=survey).order_by('-submitted_at')
    nps_questions = survey.questions.filter(question_type='nps')
    
    for response in responses:
        nps_answer = None
        comment_answer = None
        
        for answer in response.answers.all():
            if answer.question in nps_questions:
                nps_answer = answer.answer_value
            elif answer.question.question_type == 'text':
                comment_answer = answer.answer_text
        
        ws_responses.append([
            response.submitted_at.strftime('%d/%m/%Y %H:%M'),
            response.respondent_email or response.respondent_id,
            nps_questions.first().question_text if nps_questions.exists() else 'NPS',
            nps_answer or '',
            comment_answer or ''
        ])
    
    # Ajustar largura das colunas
    for col_num, column in enumerate(ws_responses.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_responses.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"NPS_{survey.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_nps_csv(request, survey_id):
    """Exporta relatório NPS em CSV"""
    survey = get_object_or_404(Survey, id=survey_id)
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"NPS_{survey.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Cabeçalhos
    writer.writerow(['Período', 'Total Respostas', 'Promotores', 'Neutros', 'Detratores', 'NPS Score'])
    
    # Dados
    results = NPSResult.objects.filter(survey=survey).order_by('-period_start', '-period_end')
    for result in results:
        writer.writerow([
            f"{result.period_start.strftime('%d/%m/%Y')} a {result.period_end.strftime('%d/%m/%Y')}",
            result.total_responses,
            result.promoters,
            result.passives,
            result.detractors,
            float(result.nps_score)
        ])
    
    return response

